import openpyxl

def getRow(file,sheetname):
    worksheet = openpyxl.load_workbook(file)
    sheet = worksheet[sheetname]
    return(sheet.max_row)

def getColumnCount(file,sheetname):
    worksheet = openpyxl.load_workbook(file)
    sheet = worksheet[sheetname]
    return(sheet.max_column)

def readData(file,sheetname,rowno,column):
    worksheet = openpyxl.load_workbook(file)
    sheet = worksheet[sheetname]
    return sheet.cell(row=rowno, column=column).value

def writeData(file, sheetname, rowno, column, data):
    worksheet = openpyxl.load_workbook(file)
    sheet = worksheet[sheetname]
    sheet.cell(row=rowno, column=column).value = data
    worksheet.save(file)